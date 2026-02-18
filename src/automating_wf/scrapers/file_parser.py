from __future__ import annotations

import csv
import re
import unicodedata
from pathlib import Path
from typing import Any


class ExportParseError(RuntimeError):
    """Raised when a trends/PinClicks export file cannot be parsed."""


HEADER_HINTS = (
    "keyword",
    "trend",
    "tendance",
    "search term",
    "search query",
    "query",
    "term",
    "topic",
    "rang",
    "rank",
    "variation",
    "change",
    "pin",
    "title",
    "description",
    "tag",
    "url",
    "link",
)


def coerce_numeric(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().lower()
    if not text:
        return 0.0

    text = text.replace(",", "")
    percent = False
    if text.endswith("%"):
        percent = True
        text = text[:-1].strip()

    multiplier = 1.0
    if text.endswith("k"):
        multiplier = 1_000.0
        text = text[:-1].strip()
    elif text.endswith("m"):
        multiplier = 1_000_000.0
        text = text[:-1].strip()
    elif text.endswith("b"):
        multiplier = 1_000_000_000.0
        text = text[:-1].strip()

    # Handles "12.4x" style values by removing trailing non-numeric chars.
    text = re.sub(r"[^0-9.\-]+$", "", text)
    if not text:
        return 0.0
    try:
        numeric = float(text) * multiplier
    except ValueError:
        return 0.0
    if percent:
        return numeric / 100.0
    return numeric


def _normalize_header_token(value: Any) -> str:
    raw = str(value or "").strip().casefold()
    normalized = unicodedata.normalize("NFKD", raw)
    ascii_only = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", ascii_only).strip()


def _is_numeric_like(value: str) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    compact = (
        text.replace("\u00a0", "")
        .replace("\u202f", "")
        .replace(" ", "")
        .replace(",", "")
    )
    return bool(re.fullmatch(r"[-+]?\d+(?:[.]\d+)?%?[kmb]?", compact))


def _detect_header_index(raw_rows: list[list[Any]]) -> int:
    best_index = 0
    best_score = float("-inf")
    hint_norm = tuple(_normalize_header_token(item) for item in HEADER_HINTS)

    for index, row in enumerate(raw_rows):
        cleaned = [str(cell or "").strip() for cell in row if str(cell or "").strip()]
        if len(cleaned) < 2:
            continue

        normalized = [_normalize_header_token(cell) for cell in cleaned]
        alpha_count = sum(1 for value in normalized if re.search(r"[a-z]", value))
        numeric_count = sum(1 for cell in cleaned if _is_numeric_like(cell))
        hint_count = 0
        for value in normalized:
            if any(hint and hint in value for hint in hint_norm):
                hint_count += 1

        score = (len(cleaned) * 3) + (alpha_count * 2) + (hint_count * 6) - (numeric_count * 3)
        if score > best_score:
            best_score = score
            best_index = index

    return best_index


def _rows_to_dicts(raw_rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not raw_rows:
        return []

    header_index = _detect_header_index(raw_rows)
    headers = [str(item or "").strip() for item in raw_rows[header_index]]
    if not any(headers):
        return []

    rows: list[dict[str, Any]] = []
    for values in raw_rows[header_index + 1 :]:
        if values is None:
            continue
        cleaned_values = [str(item or "").strip() for item in values]
        if not any(cleaned_values):
            continue
        row: dict[str, Any] = {}
        for col_index, header in enumerate(headers):
            if not header:
                continue
            row[header] = cleaned_values[col_index] if col_index < len(cleaned_values) else ""
        if len(cleaned_values) > len(headers):
            extras = [item for item in cleaned_values[len(headers) :] if item]
            if extras:
                row["None"] = extras
        if row:
            rows.append(row)
    return rows


def _parse_csv(path: Path) -> list[dict[str, Any]]:
    encodings = ("utf-8-sig", "utf-8", "cp1252")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as file_handle:
                reader = csv.reader(file_handle)
                raw_rows = [list(row) for row in reader]
                return _rows_to_dicts(raw_rows)
        except Exception as exc:
            last_error = exc
            continue
    raise ExportParseError(f"Failed to parse CSV export '{path}': {last_error}")


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExportParseError(
            "openpyxl is required to parse XLSX exports. Install dependencies."
        ) from exc

    try:
        workbook = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as exc:
        raise ExportParseError(f"Failed to open XLSX export '{path}': {exc}") from exc

    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        raw_rows = [list(values) if values is not None else [] for values in rows_iter]
        if not raw_rows:
            return []
        return _rows_to_dicts(raw_rows)
    finally:
        workbook.close()


def parse_tabular_export(path: Path) -> list[dict[str, Any]]:
    file_path = Path(path)
    if not file_path.exists():
        raise ExportParseError(f"Export file not found: {file_path}")
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return _parse_csv(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(file_path)
    raise ExportParseError(
        f"Unsupported export extension '{suffix}' for file '{file_path}'. "
        "Supported formats are CSV and XLSX."
    )
